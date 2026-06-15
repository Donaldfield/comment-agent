"""Multi-sheet Excel report generator using openpyxl.

Produces a 5-sheet .xlsx workbook:
1. Overview — Summary metrics + daily breakdown
2. Pain Points — Categorized issues with details
3. Sentiment Timeline — Daily sentiment with anomaly flags
4. Review Clusters — FAISS cluster summaries
5. Alert Log — Chronological alert history
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

from src.models.analysis import AnalysisResultBundle

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
INFO_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_report(
    bundle: AnalysisResultBundle,
    output_path: str,
) -> None:
    """Generate a 5-sheet Excel report from analysis results.

    Args:
        bundle: Complete analysis result bundle.
        output_path: Path to write .xlsx file.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _create_overview_sheet(wb, bundle)
    _create_pain_points_sheet(wb, bundle)
    _create_sentiment_timeline_sheet(wb, bundle)
    _create_clusters_sheet(wb, bundle)
    _create_alert_log_sheet(wb, bundle)

    # Ensure output directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Report saved to %s", output_path)


def _style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Apply data styling to a range of rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Estimate width for Chinese characters (~2x width)
                val = str(cell.value)
                char_len = 0
                for ch in val:
                    char_len += 2 if ord(ch) > 127 else 1
                max_len = max(max_len, char_len)
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def _create_overview_sheet(wb: Workbook, bundle: AnalysisResultBundle) -> None:
    """Sheet 1: Overview with summary metrics and daily breakdown."""
    ws = wb.create_sheet("Overview")

    total = bundle.total_reviews
    valid = bundle.valid_reviews
    positive = bundle.positive_count
    neutral = bundle.neutral_count
    negative = bundle.negative_count

    # Summary section
    summary_data = [
        ["Metric", "Value"],
        ["Product ID", bundle.product_id or "All Products"],
        ["Analysis Period", bundle.analysis_time.strftime("%Y-%m-%d %H:%M")],
        ["Total Reviews", total],
        ["Valid Reviews", f"{valid} ({_pct(valid, total)})"],
        ["Positive Reviews", f"{positive} ({_pct(positive, valid)})"],
        ["Neutral Reviews", f"{neutral} ({_pct(neutral, valid)})"],
        ["Negative Reviews", f"{negative} ({_pct(negative, valid)})"],
        [
            "Overall Sentiment Score",
            f"{_sentiment_score(positive, neutral, negative):.1f} / 5.0",
        ],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _style_header_row(ws, 2)
    _style_data_rows(ws, 2, len(summary_data), 2)

    # Daily breakdown section
    if bundle.daily_stats:
        start_row = len(summary_data) + 3
        ws.cell(row=start_row, column=1, value="Daily Sentiment Breakdown").font = Font(
            name="Microsoft YaHei", bold=True, size=12
        )

        daily_header = ["Date", "Total", "Positive", "Neutral", "Negative", "Negative %"]
        daily_start = start_row + 1
        for col_idx, header in enumerate(daily_header, 1):
            ws.cell(row=daily_start, column=col_idx, value=header)

        for i, day in enumerate(bundle.daily_stats):
            row = daily_start + 1 + i
            ws.cell(row=row, column=1, value=str(day.get("date", "")))
            ws.cell(row=row, column=2, value=day.get("total", 0))
            ws.cell(row=row, column=3, value=day.get("positive", 0))
            ws.cell(row=row, column=4, value=day.get("neutral", 0))
            ws.cell(row=row, column=5, value=day.get("negative", 0))
            total_day = max(day.get("total", 1), 1)
            ws.cell(row=row, column=6, value=f"{day.get('negative', 0) / total_day:.1%}")

        _style_header_row_custom(ws, daily_start, len(daily_header))
        _style_data_rows(ws, daily_start + 1, daily_start + len(bundle.daily_stats), len(daily_header))

        # Add a bar chart for daily sentiment
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Daily Review Sentiment"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Date"
            chart.style = 10

            data_ref = Reference(
                ws,
                min_col=2, max_col=5,
                min_row=daily_start, max_row=daily_start + len(bundle.daily_stats),
            )
            cats_ref = Reference(
                ws,
                min_col=1,
                min_row=daily_start + 1, max_row=daily_start + len(bundle.daily_stats),
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 24
            chart.height = 12

            chart_row = daily_start + len(bundle.daily_stats) + 3
            ws.add_chart(chart, f"A{chart_row}")
        except Exception as e:
            logger.warning("Chart creation failed: %s", e)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_pain_points_sheet(wb: Workbook, bundle: AnalysisResultBundle) -> None:
    """Sheet 2: Pain Points detail."""
    ws = wb.create_sheet("Pain Points")

    headers = ["Category", "Description", "Frequency", "Severity", "High Frequency", "Trend", "Example Review IDs"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, len(headers))

    for i, pp in enumerate(bundle.pain_points):
        row = i + 2
        ws.cell(row=row, column=1, value=pp.category)
        ws.cell(row=row, column=2, value=pp.description)
        ws.cell(row=row, column=3, value=pp.frequency)
        ws.cell(row=row, column=4, value=pp.severity)
        ws.cell(row=row, column=5, value="Yes" if pp.is_high_frequency else "No")
        ws.cell(row=row, column=6, value=pp.sentiment_trend)
        ws.cell(row=row, column=7, value=", ".join(pp.example_review_ids[:3]))

        # Color severity
        severity_colors = {
            "high": "FFC7CE",
            "medium": "FFEB9C",
            "low": "C6EFCE",
        }
        if pp.severity in severity_colors:
            ws.cell(row=row, column=4).fill = PatternFill(
                start_color=severity_colors[pp.severity],
                end_color=severity_colors[pp.severity],
                fill_type="solid",
            )

    _style_data_rows(ws, 2, 1 + len(bundle.pain_points), len(headers))
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_sentiment_timeline_sheet(wb: Workbook, bundle: AnalysisResultBundle) -> None:
    """Sheet 3: Sentiment Timeline with anomaly flags."""
    ws = wb.create_sheet("Sentiment Timeline")

    headers = ["Date", "Total", "Positive", "Neutral", "Negative", "Negative Ratio", "Anomaly"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, len(headers))

    anomaly_dates = set()
    for event in bundle.anomaly_events:
        if event.event_type == "sentiment_shift":
            anomaly_dates.add(event.timestamp.strftime("%Y-%m-%d"))

    for i, day in enumerate(bundle.daily_stats):
        row = i + 2
        date_str = str(day.get("date", ""))
        total = day.get("total", 0)
        negative = day.get("negative", 0)
        ratio = negative / max(total, 1)

        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=day.get("positive", 0))
        ws.cell(row=row, column=4, value=day.get("neutral", 0))
        ws.cell(row=row, column=5, value=negative)
        ws.cell(row=row, column=6, value=f"{ratio:.1%}")

        is_anomaly = date_str[:10] in anomaly_dates
        ws.cell(row=row, column=7, value="⚠ YES" if is_anomaly else "No")
        if is_anomaly:
            ws.cell(row=row, column=7).fill = CRITICAL_FILL

    _style_data_rows(ws, 2, 1 + len(bundle.daily_stats), len(headers))
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_clusters_sheet(wb: Workbook, bundle: AnalysisResultBundle) -> None:
    """Sheet 4: Review Clusters from FAISS."""
    ws = wb.create_sheet("Review Clusters")

    headers = ["Cluster ID", "Size", "Top Keywords", "Representative Review", "Avg Sentiment"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, len(headers))

    for i, cluster in enumerate(bundle.cluster_summaries):
        row = i + 2
        ws.cell(row=row, column=1, value=cluster.get("cluster_id", i))
        ws.cell(row=row, column=2, value=cluster.get("size", 0))
        ws.cell(row=row, column=3, value=", ".join(cluster.get("top_keywords", [])))
        ws.cell(row=row, column=4, value=cluster.get("representative_review", ""))
        ws.cell(row=row, column=5, value=cluster.get("avg_sentiment", ""))

    _style_data_rows(ws, 2, 1 + len(bundle.cluster_summaries), len(headers))
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _create_alert_log_sheet(wb: Workbook, bundle: AnalysisResultBundle) -> None:
    """Sheet 5: Alert Log."""
    ws = wb.create_sheet("Alert Log")

    headers = ["Alert ID", "Timestamp", "Level", "Rule", "Title", "Detail", "Remediation Plan"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, len(headers))

    for i, alert in enumerate(bundle.alerts):
        row = i + 2
        ws.cell(row=row, column=1, value=alert.id[:8])
        ws.cell(row=row, column=2, value=alert.triggered_at.strftime("%Y-%m-%d %H:%M") if hasattr(alert, 'triggered_at') else "")
        ws.cell(row=row, column=3, value=alert.level)
        ws.cell(row=row, column=4, value=alert.rule_name)
        ws.cell(row=row, column=5, value=alert.title)
        ws.cell(row=row, column=6, value=alert.detail)
        ws.cell(row=row, column=7, value=alert.remediation_plan[:500] if alert.remediation_plan else "")

        # Color by level
        level_fills = {
            "critical": CRITICAL_FILL,
            "warning": WARNING_FILL,
            "info": INFO_FILL,
        }
        if alert.level in level_fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = level_fills[alert.level]

    _style_data_rows(ws, 2, 1 + len(bundle.alerts), len(headers))
    _auto_width(ws, max_width=60)
    ws.freeze_panes = "A2"


def _style_header_row_custom(ws, row: int, num_cols: int) -> None:
    """Apply header styling to a specific row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _pct(part: int, total: int) -> str:
    """Format a percentage string."""
    if total == 0:
        return "0.0%"
    return f"{part / total:.1%}"


def _sentiment_score(positive: int, neutral: int, negative: int) -> float:
    """Calculate weighted sentiment score (1.0 - 5.0)."""
    total = positive + neutral + negative
    if total == 0:
        return 3.0
    return (positive * 5 + neutral * 3 + negative * 1) / total
